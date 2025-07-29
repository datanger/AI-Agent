import os
import time
import logging
import sys
import fcntl
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from typing import Dict, Optional, List, Set
import inotify.adapters  # For Linux file monitoring
import threading
from queue import Queue, Empty

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_monitor.log'),
        logging.StreamHandler()
    ]
)

class ExcelMonitor:
    def __init__(self, file_paths: List[str] = None):
        """
        初始化Excel监控器
        :param file_paths: 要监控的Excel文件路径列表
        """
        self.file_paths = [os.path.abspath(path) for path in file_paths] if file_paths else []
        self.excel_states: Dict[str, Dict[str, Dict[str, str]]] = {} # 用于存储Excel文件的状态
        self.notifier = inotify.adapters.Inotify()
        self.running = True
        self.handled_files = set()
        self.lock = threading.Lock()
        self.event_queue = Queue()
        self.TRIGGER_TEXT = "触发"
        self.GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
    def is_file_open(self, filepath: str) -> bool:
        """检查文件是否被其他进程打开"""
        if not os.path.exists(filepath):
            return False
            
        try:
            # 尝试以独占模式打开文件
            with open(filepath, 'a+') as f:
                fcntl.flock(f.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                fcntl.flock(f.fileno(), fcntl.LOCK_UN)
                return False
        except (IOError, BlockingIOError):
            return True
        except Exception as e:
            logging.error(f"检查文件 {filepath} 状态时出错: {str(e)}")
            return False

    def start_monitoring(self):
        """开始监控文件变化"""
        # 检查文件是否存在并加载初始状态
        for file_path in self.file_paths:
            if not os.path.exists(file_path):
                logging.error(f"文件不存在: {file_path}")
                return
            else:
                # 加载初始状态
                self.excel_states[file_path] = self._get_excel_state(file_path)
                logging.info(f"已加载初始状态: {os.path.basename(file_path)}")

        # 启动事件处理线程
        self.event_thread = threading.Thread(target=self._process_events, daemon=True)
        self.event_thread.start()

        logging.info("文件监控已启动，按 Ctrl+C 停止")

        try:
            # 为被监控的文件添加watch
            for path in self.file_paths:
                self.notifier.add_watch(path)

            for event in self.notifier.event_gen():
                if not self.running:
                    break
                if event is not None:
                    # 将事件放入队列进行处理
                    _, type_names, path, filename = event
                    file_path = os.path.join(path.decode('utf-8'), filename.decode('utf-8'))
                    if file_path in self.file_paths and ('IN_MODIFY' in type_names or 'IN_CLOSE_WRITE' in type_names):
                        self.event_queue.put(file_path)
        except KeyboardInterrupt:
            logging.info("正在停止监控...")
            self.stop()

    def _process_events(self):
        """处理文件变化事件"""
        processed = set()
        while self.running:
            try:
                file_path = self.event_queue.get(timeout=1)
                if file_path not in processed:
                    self.process_excel(file_path)
                    processed.add(file_path)
                # Clear the queue of same file path events
                while not self.event_queue.empty():
                    if self.event_queue.queue[0] == file_path:
                        self.event_queue.get()
                    else:
                        break
                processed.discard(file_path)

            except Empty:
                processed.clear()
                continue
            except Exception as e:
                logging.error(f"处理事件时出错: {str(e)}")

    def stop(self):
        """停止监控"""
        self.running = False
        logging.info("文件监控已停止")

    def _get_excel_state(self, file_path: str) -> Dict[str, Dict[str, str]]:
        """
        获取Excel文件的当前状态，表达为每个非空单元格的字典。
        :param file_path: Excel文件路径
        :return: 包含每个工作表状态的字典
        """
        state = {}
        if not os.path.exists(file_path):
            return state
        try:
            # 使用 read_only 和 data_only 模式以避免不必要的锁定和内存占用，并获取单元格的计算值
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                sheet_state = {}
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and str(cell.value).strip() != "":
                            sheet_state[cell.coordinate] = str(cell.value)
                if sheet_state:
                    state[ws_name] = sheet_state
            wb.close()
        except Exception as e:
            logging.error(f"获取 {os.path.basename(file_path)} 状态时出错: {e}")
        return state

    def _log_state_diff(self, file_path: str, old_state: Dict[str, Dict[str, str]], new_state: Dict[str, Dict[str, str]]):
        """
        比较两个状态字典，并记录变化的“事件”。
        :param file_path: 文件路径，用于日志记录
        :param old_state: 变化前的状态
        :param new_state: 变化后的状态
        """
        base_name = os.path.basename(file_path)
        
        # 如果旧状态为空，说明是第一次加载，不记录差异
        if not old_state:
            return

        # 检查工作表的变化
        old_sheets = set(old_state.keys())
        new_sheets = set(new_state.keys())

        added_sheets = new_sheets - old_sheets
        for sheet in added_sheets:
            logging.info(f"[{base_name}] Event: Sheet created '{sheet}'")

        removed_sheets = old_sheets - new_sheets
        for sheet in removed_sheets:
            logging.info(f"[{base_name}] Event: Sheet removed '{sheet}'")

        # 检查共同工作表中单元格的变化
        for sheet in old_sheets & new_sheets:
            old_sheet_state = old_state.get(sheet, {})
            new_sheet_state = new_state.get(sheet, {})
            all_cells = set(old_sheet_state.keys()) | set(new_sheet_state.keys())
            
            for cell in sorted(all_cells):
                old_value = old_sheet_state.get(cell)
                new_value = new_sheet_state.get(cell)

                if old_value != new_value:
                    if old_value is None:
                        logging.info(f"[{base_name}][{sheet}] Event: Cell created {cell} = '{new_value}'")
                    elif new_value is None:
                        logging.info(f"[{base_name}][{sheet}] Event: Cell deleted {cell} (was '{old_value}')")
                    else:
                        logging.info(f"[{base_name}][{sheet}] Event: Cell updated {cell} from '{old_value}' to '{new_value}'")

    def process_excel(self, file_path: str):
        """处理Excel文件"""
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            return
            
        # 检查文件是否被其他进程打开
        if self.is_file_open(file_path):
            logging.info(f"文件 {os.path.basename(file_path)} 正被占用，稍后重试...")
            return
            
        # 使用锁确保同一时间只有一个线程处理文件
        with self.lock:
            try:
                logging.info(f"检测到文件变化: {file_path}")
                
                # --- 状态比对和事件日志 ---
                new_state = self._get_excel_state(file_path)
                old_state = self.excel_states.get(file_path, {})
                self._log_state_diff(file_path, old_state, new_state)
                self.excel_states[file_path] = new_state # 更新状态
                
                # --- 自动高亮处理 ---
                temp_path = f"{file_path}.tmp"
                import shutil
                shutil.copy2(file_path, temp_path)
                
                wb = openpyxl.load_workbook(temp_path)
                modified = False
                
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and self.TRIGGER_TEXT in str(cell.value):
                                cell.fill = self.GREEN_FILL
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                modified = True
                                logging.info(f"在 {os.path.basename(file_path)} 的 {ws_name} 工作表中找到触发文本: {cell.coordinate}")
                
                if modified:
                    backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    shutil.copy2(file_path, backup_path)
                    logging.info(f"创建备份文件: {backup_path}")
                    wb.save(file_path)
                    logging.info(f"文件已自动高亮并更新: {file_path}")
                    # 更新高亮后的状态
                    self.excel_states[file_path] = self._get_excel_state(file_path)

                wb.close()
                
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                    
            except Exception as e:
                logging.error(f"处理文件 {file_path} 时出错: {str(e)}", exc_info=True)
                if 'temp_path' in locals() and os.path.exists(temp_path):
                    os.remove(temp_path)

def get_file_paths() -> List[str]:
    """获取用户输入的要监控的文件路径"""
    print("请输入要监控的Excel文件路径（每行一个，空行结束）:")
    file_paths = []
    
    while True:
        try:
            path = input().strip()
            if not path:
                break
                
            path = os.path.expanduser(path)
            path = os.path.abspath(path)
            
            if not os.path.exists(path):
                create = input(f"文件 {path} 不存在，是否创建? (y/n): ").strip().lower()
                if create == 'y':
                    try:
                        wb = openpyxl.Workbook()
                        wb.save(path)
                        print(f"已创建文件: {path}")
                    except Exception as e:
                        print(f"创建文件失败: {str(e)}")
                        continue
                else:
                    continue
            
            if path not in file_paths:
                file_paths.append(path)
                print(f"已添加: {path}")
            else:
                print(f"文件已添加: {path}")
                
        except KeyboardInterrupt:
            print("输入结束")
            break
        except Exception as e:
            print(f"输入错误: {str(e)}")
    
    return file_paths

if __name__ == "__main__":
    print("=== Excel文件变更事件监控器 ===")
    print("功能: 监控Excel文件变更，记录单元格级别的事件日志，并能根据'触发'文本自动高亮。")
    print("按 Ctrl+C 停止监控")
    
    # 使用硬编码路径进行测试，方便调试
    file_paths = ["/home/niejie/work/Code/tools/AI-Agent/test.xlsx"]
    
    if not file_paths:
        print("未指定要监控的文件，程序退出")
        sys.exit(0)
    
    print("开始监控以下文件:")
    for path in file_paths:
        print(f"- {path}")
    
    monitor = None
    try:
        monitor = ExcelMonitor(file_paths=file_paths)
        monitor.start_monitoring()
    except KeyboardInterrupt:
        print("正在停止监控...")
        if monitor:
            monitor.stop()
    except Exception as e:
        logging.error(f"程序出错: {str(e)}", exc_info=True)
    finally:
        print("程序已退出")
